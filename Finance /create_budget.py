from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# ── helpers ───────────────────────────────────────────────────────────────────
def fl(hex_color):
    return PatternFill(fill_type='solid', start_color=hex_color, end_color=hex_color)

def fn(size=10, bold=False, color='000000', italic=False):
    return Font(name='Arial', size=size, bold=bold, color=color, italic=italic)

def al(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bd():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def rh(ws, row, height):
    ws.row_dimensions[row].height = height

def c(ws, row, col, value='', font=None, fill=None, align=None, border=None, nf=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if nf:     cell.number_format = nf
    return cell

def mg(ws, r1, c1, r2, c2, value='', font=None, fill=None, align=None, border=None, nf=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if nf:     cell.number_format = nf
    return cell

def apply_border_range(ws, r1, c1, r2, c2):
    for row in range(r1, r2+1):
        for col in range(c1, c2+1):
            ws.cell(row=row, column=col).border = bd()

# ── colors ────────────────────────────────────────────────────────────────────
NAVY      = "1A3C6E"
BLUE      = "2E75B6"
LT_BLUE   = "BDD7EE"
PLT_BLUE  = "DEEAF1"
DK_GREEN  = "1E5631"
MD_GREEN  = "375623"
LT_GREEN  = "E2EFDA"
DK_ORANGE = "B7410E"
LT_ORANGE = "FCE4D6"
PURPLE    = "5B2C6F"
LT_PURPLE = "E8DAEF"
LT_YELLOW = "FFF2CC"
GRAY      = "595959"
LT_GRAY   = "F2F2F2"
WHITE     = "FFFFFF"
BLACK     = "000000"
BLU_IN    = "0000FF"   # blue text = manual input

IDR = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – RINGKASAN
# Referensi:
#   D8  = 'LANGGANAN TAHUNAN'!D11 / 12   (total annual di baris 11)
#   D9  = 'LANGGANAN BULANAN'!D10         (total monthly di baris 10)
#   D10 = 'IKLAN DIGITAL'!D12             (total ads di baris 12)
#   D11 = 'KEBUTUHAN LAINNYA'!E13         (total other di E13)
#   D12 = SUM(D8:D11)
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("RINGKASAN")
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = NAVY

cw(ws1,1,3); cw(ws1,2,35); cw(ws1,3,25); cw(ws1,4,26); cw(ws1,5,22); cw(ws1,6,3)

rh(ws1,1,8)
mg(ws1,2,2,2,5,"DIGITAL MARKETING BUDGET REPORT",
   font=fn(16,True,WHITE), fill=fl(NAVY), align=al('center','center'))
rh(ws1,2,36)
mg(ws1,3,2,3,5,"Periode: April 2026  |  Tim: Digital Marketing",
   font=fn(10,False,WHITE), fill=fl(BLUE), align=al('center','center'))
rh(ws1,3,20)
mg(ws1,4,2,4,5,"", fill=fl(NAVY)); rh(ws1,4,5)
rh(ws1,5,12)

mg(ws1,6,2,6,5,"RINGKASAN ESTIMASI BIAYA BULANAN",
   font=fn(11,True,WHITE), fill=fl(BLUE), align=al('center','center'))
rh(ws1,6,22)

hdrs = ["KATEGORI","KETERANGAN","ESTIMASI BIAYA (Rp)","CATATAN"]
for i,txt in enumerate(hdrs,2):
    c(ws1,7,i,txt, font=fn(10,True,WHITE), fill=fl(NAVY), align=al('center','center'), border=bd())
rh(ws1,7,20)

summary_rows = [
    # row 8
    ("Langganan Tahunan  (setara/bulan)",
     "Envato, Canva",
     "='LANGGANAN TAHUNAN'!D11/12",
     "Biaya tahunan dibagi 12"),
    # row 9
    ("Langganan Bulanan",
     "ChatGPT, Domain, Google Workspace",
     "='LANGGANAN BULANAN'!D10",
     ""),
    # row 10
    ("Budget Iklan Digital",
     "Meta Ads, Google Ads, TikTok Ads",
     "='IKLAN DIGITAL'!D12",
     "Total Rp 16.000.000"),
    # row 11
    ("Kebutuhan Lainnya",
     "Kolaborasi, Giveaway, Peralatan",
     "='KEBUTUHAN LAINNYA'!E13",
     ""),
]
for i,(cat,ket,amt,note) in enumerate(summary_rows, 8):
    rf = fl(PLT_BLUE) if i%2==0 else fl(WHITE)
    c(ws1,i,2,cat,  font=fn(10,False,BLACK), fill=rf, align=al('left','center',True), border=bd())
    c(ws1,i,3,ket,  font=fn(10,False,BLACK), fill=rf, align=al('left','center'),      border=bd())
    c(ws1,i,4,amt,  font=fn(10,False,BLACK), fill=rf, align=al('right','center'),     border=bd(), nf=IDR)
    c(ws1,i,5,note, font=fn(9,False,GRAY,True), fill=rf, align=al('left','center',True), border=bd())
    rh(ws1,i,20)

# total
mg(ws1,12,2,12,3,"TOTAL ESTIMASI BIAYA BULANAN",
   font=fn(11,True,WHITE), fill=fl(NAVY), align=al('center','center'))
ws1.cell(row=12,column=2).border = bd()
ws1.cell(row=12,column=3).border = bd()
c(ws1,12,4,"=SUM(D8:D11)", font=fn(11,True,WHITE), fill=fl(NAVY),
  align=al('right','center'), border=bd(), nf=IDR)
c(ws1,12,5,"", fill=fl(NAVY), border=bd())
rh(ws1,12,26)

rh(ws1,13,12)

# legend
mg(ws1,14,2,14,5,"KETERANGAN WARNA:",
   font=fn(10,True,BLACK), align=al('left','center'))
rh(ws1,14,18)
legend = [
    (15, "● Teks Biru  = Input manual (angka yang dapat diubah sesuai data aktual)", BLU_IN, WHITE),
    (16, "● Teks Hitam = Formula otomatis (jangan diubah secara manual)",            BLACK,  WHITE),
    (17, "● Latar Kuning = Sel yang perlu diisi / diverifikasi",                     GRAY,   LT_YELLOW),
]
for row,txt,col,bg in legend:
    mg(ws1,row,2,row,5,txt, font=fn(9,False,col), align=al('left','center'))
    for col_idx in range(2,6):
        ws1.cell(row=row,column=col_idx).fill = fl(bg)
    rh(ws1,row,16)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – LANGGANAN TAHUNAN
# Total annual cost di D11 = SUM(D9:D10)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("LANGGANAN TAHUNAN")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = DK_GREEN

cw(ws2,1,3); cw(ws2,2,28); cw(ws2,3,25); cw(ws2,4,24); cw(ws2,5,24); cw(ws2,6,22); cw(ws2,7,3)

rh(ws2,1,8)
mg(ws2,2,2,2,6,"LANGGANAN TAHUNAN  (Annual Subscriptions)",
   font=fn(14,True,WHITE), fill=fl(DK_GREEN), align=al('center','center'))
rh(ws2,2,32)
mg(ws2,3,2,3,6,"Periode: 2026  |  Tim: Digital Marketing",
   font=fn(10,False,WHITE), fill=fl(MD_GREEN), align=al('center','center'))
rh(ws2,3,20)
mg(ws2,4,2,4,6,"", fill=fl(DK_GREEN)); rh(ws2,4,5)
rh(ws2,5,10)
mg(ws2,6,2,6,6,
   "Catatan: Biaya tahunan dibayar satu kali dalam setahun. "
   "Kolom 'Setara/Bulan' digunakan untuk alokasi biaya bulanan pada laporan ringkasan.",
   font=fn(9,False,GRAY,True), fill=fl(LT_GREEN), align=al('left','center',True))
rh(ws2,6,32)
rh(ws2,7,10)

hdrs2 = ["NAMA TOOLS","KATEGORI / FUNGSI","BIAYA TAHUNAN (Rp)","SETARA / BULAN (Rp)","CATATAN"]
for i,txt in enumerate(hdrs2,2):
    c(ws2,8,i,txt, font=fn(10,True,WHITE), fill=fl(DK_GREEN),
      align=al('center','center',True), border=bd())
rh(ws2,8,22)

# baris 9–10 = data; baris 11 = total
annual_items = [
    (9,  "Envato Elements",  "Template Desain, Aset Video, Musik Royalti-bebas"),
    (10, "Canva Pro",        "Desain Grafis, Konten Visual, Presentasi"),
]
for row, name, func in annual_items:
    rf = fl(LT_GREEN) if row%2==0 else fl(WHITE)
    c(ws2,row,2,name, font=fn(10,True,BLACK),    fill=rf,          align=al('left','center'),      border=bd())
    c(ws2,row,3,func, font=fn(10,False,BLACK),   fill=rf,          align=al('left','center',True), border=bd())
    # D = input biaya tahunan (biru + latar kuning)
    c(ws2,row,4,"",   font=fn(10,False,BLU_IN),  fill=fl(LT_YELLOW), align=al('right','center'),  border=bd(), nf=IDR)
    # E = D/12 formula
    c(ws2,row,5,f"=IF(D{row}>0,D{row}/12,\"-\")",
                      font=fn(10,False,BLACK),   fill=rf,          align=al('right','center'),     border=bd(), nf=IDR)
    c(ws2,row,6,"",   font=fn(9,False,GRAY,True),fill=rf,          align=al('left','center',True), border=bd())
    rh(ws2,row,22)

# total baris 11 ← RINGKASAN merujuk D11
mg(ws2,11,2,11,3,"TOTAL BIAYA TAHUNAN",
   font=fn(11,True,WHITE), fill=fl(DK_GREEN), align=al('center','center'))
ws2.cell(row=11,column=2).border = bd()
ws2.cell(row=11,column=3).border = bd()
c(ws2,11,4,"=SUM(D9:D10)", font=fn(11,True,WHITE), fill=fl(DK_GREEN),
  align=al('right','center'), border=bd(), nf=IDR)
c(ws2,11,5,"=SUM(E9:E10)", font=fn(11,True,WHITE), fill=fl(DK_GREEN),
  align=al('right','center'), border=bd(), nf=IDR)
c(ws2,11,6,"", fill=fl(DK_GREEN), border=bd())
rh(ws2,11,26)

rh(ws2,12,10)
mg(ws2,13,2,13,6,
   "Isilah kolom 'BIAYA TAHUNAN (Rp)' dengan harga aktual invoice/kuitansi masing-masing tools.",
   font=fn(9,False,NAVY,True), fill=fl(PLT_BLUE), align=al('left','center',True))
rh(ws2,13,28)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – LANGGANAN BULANAN
# Total monthly di D10 = SUM(D7:D9)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("LANGGANAN BULANAN")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = PURPLE

cw(ws3,1,3); cw(ws3,2,28); cw(ws3,3,25); cw(ws3,4,24); cw(ws3,5,22); cw(ws3,6,3)

rh(ws3,1,8)
mg(ws3,2,2,2,5,"LANGGANAN BULANAN  (Monthly Subscriptions)",
   font=fn(14,True,WHITE), fill=fl(PURPLE), align=al('center','center'))
rh(ws3,2,32)
mg(ws3,3,2,3,5,"Periode: April 2026  |  Tim: Digital Marketing",
   font=fn(10,False,WHITE), fill=fl("7D3C98"), align=al('center','center'))
rh(ws3,3,20)
mg(ws3,4,2,4,5,"", fill=fl(PURPLE)); rh(ws3,4,5)
rh(ws3,5,12)

hdrs3 = ["NAMA TOOLS","KATEGORI / FUNGSI","BIAYA BULANAN (Rp)","CATATAN"]
for i,txt in enumerate(hdrs3,2):
    c(ws3,6,i,txt, font=fn(10,True,WHITE), fill=fl(PURPLE),
      align=al('center','center',True), border=bd())
rh(ws3,6,20)

# baris 7–9 = data; baris 10 = total
monthly_items = [
    (7,  "ChatGPT (OpenAI)",   "AI Tools – Copywriting, Riset Konten, Ide Kreatif", "ChatGPT Plus / Team Plan"),
    (8,  "Domain",             "Hosting Website / Landing Page Marketing",           "Domain .com / .id per tahun → dibagi 12"),
    (9,  "Google Workspace",   "Email Bisnis, Google Drive, Docs, Sheets",           "Per user / bulan"),
]
for row, name, func, note in monthly_items:
    rf = fl(LT_PURPLE) if row%2==0 else fl(WHITE)
    c(ws3,row,2,name, font=fn(10,True,BLACK),    fill=rf,           align=al('left','center'),       border=bd())
    c(ws3,row,3,func, font=fn(10,False,BLACK),   fill=rf,           align=al('left','center',True),  border=bd())
    c(ws3,row,4,"",   font=fn(10,False,BLU_IN),  fill=fl(LT_YELLOW),align=al('right','center'),      border=bd(), nf=IDR)
    c(ws3,row,5,note, font=fn(9,False,GRAY,True),fill=rf,           align=al('left','center',True),  border=bd())
    rh(ws3,row,22)

# total baris 10 ← RINGKASAN merujuk D10
mg(ws3,10,2,10,3,"TOTAL LANGGANAN BULANAN",
   font=fn(11,True,WHITE), fill=fl(PURPLE), align=al('center','center'))
ws3.cell(row=10,column=2).border = bd()
ws3.cell(row=10,column=3).border = bd()
c(ws3,10,4,"=SUM(D7:D9)", font=fn(11,True,WHITE), fill=fl(PURPLE),
  align=al('right','center'), border=bd(), nf=IDR)
c(ws3,10,5,"", fill=fl(PURPLE), border=bd())
rh(ws3,10,26)

rh(ws3,11,10)
mg(ws3,12,2,12,5,
   "Isilah kolom 'BIAYA BULANAN (Rp)' dengan tarif aktual setiap tools.",
   font=fn(9,False,NAVY,True), fill=fl(PLT_BLUE), align=al('left','center',True))
rh(ws3,12,28)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – IKLAN DIGITAL
# Layout:
#   Row 6:   Input total budget (D6 = 16.000.000)
#   Row 8:   Header
#   Row 9:   Meta Ads
#   Row 10:  Google Ads
#   Row 11:  TikTok Ads
#   Row 12:  TOTAL  ← RINGKASAN merujuk D12
# Kolom: B=Platform, C=Alokasi%, D=Budget(Rp), E=CPL(Rp), F=Est.Leads, G=Catatan
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("IKLAN DIGITAL")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = DK_ORANGE

cw(ws4,1,3); cw(ws4,2,22); cw(ws4,3,18); cw(ws4,4,22); cw(ws4,5,20)
cw(ws4,6,20); cw(ws4,7,28); cw(ws4,8,3)

rh(ws4,1,8)
mg(ws4,2,2,2,7,"BUDGET IKLAN DIGITAL  (Digital Ads Budget)",
   font=fn(14,True,WHITE), fill=fl(DK_ORANGE), align=al('center','center'))
rh(ws4,2,32)
mg(ws4,3,2,3,7,"Periode: April 2026  |  Tim: Digital Marketing",
   font=fn(10,False,WHITE), fill=fl("D35400"), align=al('center','center'))
rh(ws4,3,20)
mg(ws4,4,2,4,7,"", fill=fl(DK_ORANGE)); rh(ws4,4,5)
rh(ws4,5,12)

# total budget input baris 6
mg(ws4,6,2,6,3,"TOTAL BUDGET IKLAN BULAN INI  (Rp):",
   font=fn(10,True,BLACK), fill=fl(LT_ORANGE), align=al('left','center'))
ws4.cell(row=6,column=2).border = bd()
ws4.cell(row=6,column=3).border = bd()
c(ws4,6,4,16000000, font=fn(12,True,BLU_IN), fill=fl(LT_YELLOW),
  align=al('right','center'), border=bd(), nf=IDR)
mg(ws4,6,5,6,7,"← Input total budget (dapat diubah sesuai bulan berjalan)",
   font=fn(9,False,GRAY,True), fill=fl(LT_ORANGE), align=al('left','center'))
apply_border_range(ws4,6,5,6,7)
rh(ws4,6,26)
rh(ws4,7,12)

# header baris 8
hdrs4 = ["PLATFORM","ALOKASI (%)","BUDGET (Rp)","CPL – Cost/Lead (Rp)","EST. LEADS","TUJUAN / CATATAN"]
for i,txt in enumerate(hdrs4,2):
    c(ws4,8,i,txt, font=fn(10,True,WHITE), fill=fl(DK_ORANGE),
      align=al('center','center',True), border=bd())
rh(ws4,8,22)

# data baris 9–11
#   F (col 6) = Estimasi Leads = D / E   (Budget / CPL)
ads = [
    (9,  "Meta Ads",   0.60, 178000, f"=IF(E9>0,D9/E9,\"-\")",  "Lead Generation & Remarketing"),
    (10, "Google Ads", 0.30, 45000,  f"=IF(E10>0,D10/E10,\"-\")", "Lead Generation – Search Intent"),
    (11, "TikTok Ads", 0.10, None,   "Awareness",                 "Brand Awareness – tanpa target CPL"),
]
for row, plat, pct, cpl_val, leads_val, note in ads:
    rf = fl(LT_ORANGE) if row%2==0 else fl(WHITE)
    c(ws4,row,2,plat,    font=fn(10,True,BLACK),   fill=rf,           align=al('left','center'),       border=bd())
    # C = alokasi % – input
    c(ws4,row,3,pct,     font=fn(11,True,BLU_IN),  fill=fl(LT_YELLOW),align=al('center','center'),     border=bd(), nf=PCT)
    # D = budget formula
    c(ws4,row,4,f"=$D$6*C{row}", font=fn(10,False,BLACK), fill=rf,   align=al('right','center'),       border=bd(), nf=IDR)
    # E = CPL – input (atau dash)
    if cpl_val:
        c(ws4,row,5,cpl_val, font=fn(10,False,BLU_IN), fill=fl(LT_YELLOW), align=al('right','center'), border=bd(), nf=IDR)
    else:
        c(ws4,row,5,"-",     font=fn(10,False,GRAY,True), fill=rf,    align=al('center','center'),     border=bd())
    # F = estimasi leads formula / teks
    if leads_val.startswith("="):
        c(ws4,row,6,leads_val, font=fn(10,False,BLACK), fill=rf,      align=al('right','center'),      border=bd(), nf='#,##0;(#,##0);"-"')
    else:
        c(ws4,row,6,leads_val, font=fn(9,False,GRAY,True), fill=rf,   align=al('center','center'),     border=bd())
    c(ws4,row,7,note,    font=fn(9,False,GRAY,True), fill=rf,         align=al('left','center',True),  border=bd())
    rh(ws4,row,22)

# total baris 12 ← RINGKASAN merujuk D12
mg(ws4,12,2,12,2,"TOTAL",
   font=fn(11,True,WHITE), fill=fl(DK_ORANGE), align=al('center','center'))
ws4.cell(row=12,column=2).border = bd()
c(ws4,12,3,"=SUM(C9:C11)", font=fn(11,True,WHITE), fill=fl(DK_ORANGE),
  align=al('center','center'), border=bd(), nf=PCT)
c(ws4,12,4,"=SUM(D9:D11)", font=fn(11,True,WHITE), fill=fl(DK_ORANGE),
  align=al('right','center'), border=bd(), nf=IDR)
for col in [5,6,7]:
    c(ws4,12,col,"", fill=fl(DK_ORANGE), border=bd())
rh(ws4,12,26)

rh(ws4,13,10)
mg(ws4,14,2,14,7,
   "Catatan: CPL = Cost Per Lead. Estimasi Leads = Budget dibagi CPL. "
   "TikTok Ads digunakan untuk Brand Awareness, tidak memiliki target CPL.",
   font=fn(9,False,NAVY,True), fill=fl(PLT_BLUE), align=al('left','center',True))
rh(ws4,14,32)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – KEBUTUHAN LAINNYA
# Layout:
#   Row 6:  Header
#   Row 7–12: data (6 item)
#   Row 13: TOTAL ← RINGKASAN merujuk E13
# Kolom: B=Kategori, C=Item, D=Qty, E=Est.Biaya(Rp), F=Catatan
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("KEBUTUHAN LAINNYA")
ws5.sheet_view.showGridLines = False
ws5.sheet_properties.tabColor = NAVY

cw(ws5,1,3); cw(ws5,2,22); cw(ws5,3,30); cw(ws5,4,8); cw(ws5,5,24); cw(ws5,6,28); cw(ws5,7,3)

rh(ws5,1,8)
mg(ws5,2,2,2,6,"KEBUTUHAN LAINNYA  (Other Budget Needs)",
   font=fn(14,True,WHITE), fill=fl(NAVY), align=al('center','center'))
rh(ws5,2,32)
mg(ws5,3,2,3,6,"Periode: April 2026  |  Tim: Digital Marketing",
   font=fn(10,False,WHITE), fill=fl(BLUE), align=al('center','center'))
rh(ws5,3,20)
mg(ws5,4,2,4,6,"", fill=fl(NAVY)); rh(ws5,4,5)
rh(ws5,5,12)

hdrs5 = ["KATEGORI","ITEM / KETERANGAN","QTY","ESTIMASI BIAYA (Rp)","CATATAN"]
for i,txt in enumerate(hdrs5,2):
    c(ws5,6,i,txt, font=fn(10,True,WHITE), fill=fl(NAVY),
      align=al('center','center',True), border=bd())
rh(ws5,6,20)

# baris 7–12 = data; baris 13 = total
other_items = [
    (7,  "Kolaborasi", "Dreamlab – Fee Kerjasama",           1, None,   "Kerjasama dengan Klik Media – konfirmasi nilai kontrak"),
    (8,  "Kolaborasi", "Klik Media – Fee Kerjasama",         1, None,   "Konfirmasi nilai kontrak / MOU"),
    (9,  "Marketing",  "Giveaway Produk (Hadiah Konten)",    1, 300000, "Estimasi ~Rp 300.000 per item (angka dapat berubah)"),
    (10, "Marketing",  "Giveaway Karyawan (Internal)",       1, None,   "Qty & nilai sesuai keputusan manajemen"),
    (11, "Peralatan",  "Background Seamless (Selambu)",      1, None,   "Verifikasi harga aktual – cek supplier"),
    (12, "Peralatan",  "Softbox Lighting",                   1, None,   "1 unit – verifikasi harga aktual"),
]
for row, cat, item, qty, cost, note in other_items:
    rf = fl(PLT_BLUE) if row%2==0 else fl(WHITE)
    c(ws5,row,2,cat,  font=fn(10,True,BLACK),    fill=rf,           align=al('left','center'),       border=bd())
    c(ws5,row,3,item, font=fn(10,False,BLACK),   fill=rf,           align=al('left','center',True),  border=bd())
    c(ws5,row,4,qty,  font=fn(10,False,BLU_IN),  fill=fl(LT_YELLOW),align=al('center','center'),     border=bd())
    if cost:
        c(ws5,row,5,cost, font=fn(10,False,BLU_IN), fill=fl(LT_YELLOW),align=al('right','center'),   border=bd(), nf=IDR)
    else:
        c(ws5,row,5,"",   font=fn(10,False,BLU_IN), fill=fl(LT_YELLOW),align=al('right','center'),   border=bd(), nf=IDR)
    c(ws5,row,6,note, font=fn(9,False,GRAY,True),fill=rf,           align=al('left','center',True),  border=bd())
    rh(ws5,row,22)

# total baris 13 ← RINGKASAN merujuk E13
mg(ws5,13,2,13,4,"TOTAL KEBUTUHAN LAINNYA",
   font=fn(11,True,WHITE), fill=fl(NAVY), align=al('center','center'))
for col in [2,3,4]:
    ws5.cell(row=13,column=col).border = bd()
c(ws5,13,5,"=SUM(E7:E12)", font=fn(11,True,WHITE), fill=fl(NAVY),
  align=al('right','center'), border=bd(), nf=IDR)
c(ws5,13,6,"", fill=fl(NAVY), border=bd())
rh(ws5,13,26)

rh(ws5,14,10)
mg(ws5,15,2,15,6,
   "Isilah kolom 'ESTIMASI BIAYA' yang masih kosong dengan harga aktual. "
   "Harga background (selambu) dan softbox harap dikonfirmasi ke supplier.",
   font=fn(9,False,NAVY,True), fill=fl(PLT_BLUE), align=al('left','center',True))
rh(ws5,15,32)

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
out = "/Users/revitayustianawati/Desktop/Lovely Work /Budget_Digital_Marketing_April_2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
